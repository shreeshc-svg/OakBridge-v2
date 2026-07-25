"""
Regenerate `release_order.json` from the go-live Title Master.

Ordered newest -> oldest by real publication date (rank 1 = most recent), which is
exactly what `release_rank` means to the storefront: it drives the "New Arrivals"
sort and the homepage "Hot Off the Press" row.

Why this exists: the previous release_order.json was built from an earlier master
(247 rows) and applied to the catalogue BEFORE the go-live merge added 51 titles.
Those titles therefore carried no `release_rank` and could never appear as new
arrivals, however recent they were.

Titles with no publication date in the sheet are ranked LAST (they keep a rank so
they are still deterministic, but they never jump the queue).

    python backend/build_release_order.py          # rewrite release_order.json
    python backend/build_release_order.py --check  # report only, write nothing
"""
from __future__ import annotations

import argparse
import datetime
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
XLSX = HERE / "title_master_go_live.xlsx"
OUT = HERE / "release_order.json"


def build() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    recs = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[3]:
            continue
        d = r[13]
        if not isinstance(d, (datetime.datetime, datetime.date)):
            d = None
        recs.append({
            "isbn": str(r[3]).strip(),
            "title": str(r[6]).strip() if r[6] else "",
            "_date": d,
        })

    # Newest first. Undated titles sort to the very end (sentinel date.min).
    recs.sort(key=lambda x: (x["_date"] is None, -(x["_date"] or datetime.datetime.min).toordinal()))

    out = []
    for i, x in enumerate(recs, start=1):
        d = x["_date"]
        out.append({
            "rank": i,
            "isbn": x["isbn"],
            "publication_date": d.strftime("%Y-%m-%d") if d else None,
            "year": d.year if d else None,
            "month": d.month if d else None,
            "in_catalogue": True,
            "title": x["title"],
        })
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true", help="report only; do not write")
    args = ap.parse_args()

    data = build()
    dated = [d for d in data if d["publication_date"]]
    print(f"titles: {len(data)}  (dated {len(dated)}, undated {len(data) - len(dated)})")
    print("newest 5:")
    for d in data[:5]:
        print(f"  {d['rank']:>3}. {d['publication_date']}  {d['title'][:60]}")
    print("oldest dated:")
    for d in dated[-3:]:
        print(f"  {d['rank']:>3}. {d['publication_date']}  {d['title'][:60]}")

    if args.check:
        print("\n[check] nothing written.")
        return 0

    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nwrote {OUT} ({len(data)} entries)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
