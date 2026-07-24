"""
Merge the go-live Title Master into the live catalogue — WITHOUT disturbing
existing titles.

    cd backend
    python merge_titles.py --dry-run      # preview: added / existing / obsolete
    python merge_titles.py --apply        # insert the NEW titles only
    python merge_titles.py --apply --remove-obsolete   # also delete live titles
                                                        # whose ISBN isn't in the sheet

Strategy (per go-live decisions):
- Match the sheet by ISBN.
- Titles already live (ISBN match) are left EXACTLY as they are — their price,
  cover, stock, bestseller/new-release flags, rating and discount are preserved.
- Titles in the sheet but not live are INSERTED (mapped below; cover points at the
  ISBN-named file on S3; price from the sheet; sensible defaults for the rest).
- Titles live but not in the sheet are only removed if you pass --remove-obsolete.

Category mapping: Academic->academic ; Law/Tax->professional (subject Law/Tax) ;
BGR->bgr (Business & General).

Connection string: reads mongo.txt (same file as the other loaders), or MONGO_URL,
or prompts. Must point at Atlas (production).
"""
from __future__ import annotations

import argparse
import datetime
import os
import re
import sys
import uuid

import openpyxl
from dotenv import load_dotenv
from pymongo import MongoClient

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "title_master_go_live.xlsx")

CATMAP = {
    "Academic": ("academic", "Academic"),
    "Law": ("professional", "Law"),
    "Tax": ("professional", "Tax"),
    "BGR": ("bgr", "Business & General"),
}
def _binding(v):
    """Normalise HB / HB(DJ) / PB / PB(GF) / Pc variants to Hardcover / Paperback."""
    s = str(v or "").strip()
    if not s:
        return None
    u = s.upper()
    if u.startswith("HB"):
        return "Hardcover"
    if u.startswith("PB") or u.startswith("PC"):
        return "Paperback"
    return s


def _year(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.year
    try:
        return int(str(v)[:4])
    except Exception:  # noqa: BLE001
        return 2024


def _size(l, w, h):
    parts = [str(x).strip() for x in (l, w, h) if x not in (None, "")]
    return f"{l} x {w} x {h} cm" if len(parts) == 3 else None


def build_records():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    recs = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[3]:
            continue
        subj = str(r[1]).strip() if r[1] else ""
        cat, subject = CATMAP.get(subj, ("bgr", subj or "General"))
        isbn = str(r[3]).strip()
        try:
            price = float(r[15]) if r[15] not in (None, "") else 0.0
        except Exception:  # noqa: BLE001
            price = 0.0
        recs.append({
            "id": str(uuid.uuid4()),
            "title": str(r[6]).strip() if r[6] else "",
            "subtitle": None,
            "author": str(r[5]).strip() if r[5] else "",
            "author_bio": (str(r[8]).strip() if r[8] else None),
            "author_photo": None,
            "isbn": isbn,
            "category": cat,
            "subject": subject,
            "grade": None,
            "binding": _binding(r[11]),
            "size": _size(r[16], r[17], r[18]),
            "description": str(r[7]).strip() if r[7] else "",
            "price": price,
            "original_price": None,
            "cover_image": f"/api/files/oakbridge/covers/{isbn}.jpg",
            "pages": int(r[14]) if str(r[14] or "").strip().isdigit() else 0,
            "language": str(r[10]).strip() if r[10] else "English",
            "publisher": "Oakbridge Publishing",
            "publication_year": _year(r[13]),
            "bestseller": False,
            "new_release": False,
            "rating": 4.5,
            "stock": 100,
            "has_ebook": False,
            "variants": [],
        })
    return recs


def _mongo():
    env = os.path.join(HERE, ".env")
    if os.path.exists(env):
        load_dotenv(env, override=True)
    m = os.environ.get("MONGO_URL")
    txt = os.path.join(HERE, "mongo.txt")
    if not m and os.path.exists(txt):
        m = open(txt, encoding="utf-8").read().strip().strip('"').strip("'")
        print(f"read connection string from {txt}")
    if not m:
        m = input("Atlas URL (mongodb+srv://...): ").strip().strip('"').strip("'")
    if not m:
        print("ERROR: no connection string.", file=sys.stderr)
        raise SystemExit(2)
    dbname = os.environ.get("DB_NAME")
    if not dbname:
        mt = re.search(r"\.net/([^?]+)", m)
        guess = (mt.group(1) if mt else "").strip("/") or "oakbridge"
        dbname = input(f"Database name [{guess}]: ").strip() or guess
    client = MongoClient(m, serverSelectionTimeoutMS=15000)
    client.admin.command("ping")
    return client[dbname]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--remove-obsolete", action="store_true")
    args = ap.parse_args()
    if not args.apply:
        args.dry_run = True

    recs = build_records()
    sheet_isbns = {r["isbn"] for r in recs}
    print(f"sheet titles: {len(recs)} | unique ISBNs: {len(sheet_isbns)}")

    db = _mongo()
    live = list(db.books.find({}, {"_id": 0, "isbn": 1, "title": 1}))
    live_isbns = {str(b.get("isbn", "")).strip() for b in live}
    print(f"live catalogue: {len(live)} titles")

    to_add = [r for r in recs if r["isbn"] not in live_isbns]
    already = [r for r in recs if r["isbn"] in live_isbns]
    obsolete = [b for b in live if str(b.get("isbn", "")).strip() not in sheet_isbns]

    print(f"\n  NEW (in sheet, not live)      : {len(to_add)}")
    print(f"  EXISTING (kept as-is)         : {len(already)}")
    print(f"  OBSOLETE (live, not in sheet) : {len(obsolete)}")
    if to_add[:5]:
        print("\n  first new titles:")
        for r in to_add[:5]:
            print(f"    + [{r['category']}/{r['subject']}] {r['title']}  (Rs {r['price']:.0f})  {r['isbn']}")
    if obsolete:
        print("\n  obsolete (would be removed only with --remove-obsolete):")
        for b in obsolete[:20]:
            print(f"    - {b.get('title','?')}  {b.get('isbn','')}")
        if len(obsolete) > 20:
            print(f"    ... and {len(obsolete) - 20} more")

    if args.dry_run:
        print("\n[dry-run] no changes made. Re-run with --apply to insert the new titles.")
        return 0

    if to_add:
        db.books.insert_many(to_add)
        print(f"\nInserted {len(to_add)} new titles.")
    else:
        print("\nNo new titles to insert.")

    if args.remove_obsolete and obsolete:
        ids = [str(b.get("isbn", "")).strip() for b in obsolete]
        removed = db.books.delete_many({"isbn": {"$in": ids}}).deleted_count
        print(f"Removed {removed} obsolete titles.")

    print("\nDone. New titles need their cover images uploaded to S3 at "
          "oakbridge/covers/<ISBN>.jpg (same as before) or they'll show a placeholder.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
